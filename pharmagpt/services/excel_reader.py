"""
services/excel_reader.py — Extract text from Excel (.xlsx) files using openpyxl.

Excel is heavily used in pharma for:
  - OQ/PQ test data sheets
  - FMEA risk registers
  - SPC / OEE dashboards
  - Vendor qualification scorecards

Each worksheet is extracted as a labelled section so Gemini knows which
sheet the data came from when answering questions.

read_only=True avoids loading charts and images into memory.
data_only=True reads computed cell values instead of formulas.
"""

import openpyxl


def extract(file_path: str) -> tuple[str, int]:
    """
    Extract text from a .xlsx file and return (text, sheet_count).

    Each non-empty sheet becomes a labelled block:
        [Sheet: Test Results]
        Col A Header | Col B Header | ...
        Value 1      | Value 2      | ...

    Parameters
    ----------
    file_path : str — absolute path to the .xlsx file

    Returns
    -------
    text        : str — all sheet data as labelled text blocks
    sheet_count : int — number of sheets in the workbook
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_count = len(wb.sheetnames)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if cells:
                rows.append(" | ".join(cells))

        if rows:
            sections.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    wb.close()
    return "\n\n".join(sections), sheet_count
